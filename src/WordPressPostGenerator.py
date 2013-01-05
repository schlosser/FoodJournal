"""
File -- WordPressPostGenerator.py
Creator -- Dan Schlosser
Date -- 11/21/12
Description -- Generates Food Journal WordPress blog posts from photos and 
               and Excel file that contains the meal descriptions.

"""
from wordpress_xmlrpc import Client, WordPressPost
from wordpress_xmlrpc.compat import xmlrpc_client
from wordpress_xmlrpc.methods.posts import GetPosts, NewPost
from wordpress_xmlrpc.methods.users import GetUserInfo
from wordpress_xmlrpc.methods.media import UploadFile
from openpyxl.reader.excel import load_workbook
import ConfigParser
import mimetypes
import datetime
import getpass
import os
import pyexiv2

# ||-------------------------------------------------------------------------||
# ||-Main -------------------------------------------------------------------||
# ||-------------------------------------------------------------------------||

def main():
    print getOpening()
    CONFIG_DIR = ''
    CONFIG_FILENAME = 'Config.ini' 
    CONFIG_SECTION = 'WordPress_Post_Generator'
    values = {'EXCEL_FILENAME': '',
              'EXCEL_DIR': '',
              'SHEET_NAME': '',
              'PHOTOS_DIR': '',
              'WP_URL' : '',
              'WP_USERNAME': ''}
    values = getConfig(CONFIG_DIR + CONFIG_FILENAME, CONFIG_SECTION, values) 
    wp = initWordpressConnection(values['WP_URL'], values['WP_USERNAME'])
    isPostSuccessful = loopDates(wp, values['EXCEL_DIR'] + 
                                 values['EXCEL_FILENAME'], 
                                 values['SHEET_NAME'], values['PHOTOS_DIR'])
    print getClosing(isPostSuccessful),
    raw_input('  Press any key to exit:')

def loopDates(wp, EXCEL_DIR, SHEET_NAME, PHOTOS_DIR):
    """Iterates through the unposted Excel rows, and posts a WordPress post for
    each day, with photos. Returns whether or not each post was successful.
    
    Arguments:
    wp -- The authenticated WordPress client instance.
    EXCEL_DIR -- Filename of the Excel file to be read from.
    SHEET_NAME -- Name of the sheet within the Excel file.
    PHOTOS_DIR -- Directory where photos are stored.
    
    Return:
    isPostSuccessful -- Dictionary of blog post names and whether or not they 
                        were successful.
    """
    isPostSuccessful = {}  # [attempted, successful]
    latest = getLatestBlogPostDate(wp)
    (meals,thisDay) = getNextDaysMeals(latest, EXCEL_DIR, SHEET_NAME)
    
    while meals:
        if not isPostSuccessful:
            print('\nWorking'),
        uploadedFiles = uploadPictures(wp, thisDay, meals, PHOTOS_DIR)
        text = getTextBody(meals, uploadedFiles)
        (title, wasSuccessful) = postToWordpress(text, thisDay, wp)
        isPostSuccessful[title] = wasSuccessful
        latest = thisDay
        (meals, thisDay) = getNextDaysMeals(latest, EXCEL_DIR, SHEET_NAME)
    return isPostSuccessful

    
# ||-------------------------------------------------------------------------||
# ||-I/O---------------------------------------------------------------------||
# ||-------------------------------------------------------------------------||

def getConfig(CONFIG_FILENAME, CONFIG_SECTION, values):
    """Read from the configuration file, and return the values in a dictionary.
    
    Arguments:
    CONFIG_FILENAME -- Filename of the configuration file.
    CONFIG_SECTION -- Section within the configuration file from which to read.
    values -- Dictionary of variables to be filled.
    
    Return:
    values -- A dictionary of variables filled from the configuration file.
    
    """
    config = ConfigParser.SafeConfigParser(allow_no_value=True)
    config.read(CONFIG_FILENAME)
    for key in values:
        if config.get(CONFIG_SECTION, key):
            values[key] = config.get(CONFIG_SECTION, key)
    return values

def getOpening():
    """Creates a string to be printed at the beginning of the program.
    
    Return:
    opening -- The string to be printed.
    
    """
    
    opening = '''
         __      __          _ ___                        
         \ \    / /__ _ _ __| | _ \_ _ ___ ______         
          \ \/\/ / _ \ '_/ _` |  _/ '_/ -_|_-<_-<         
           \_/\_/\___/_| \__,_|_| |_| \___/__/__/         
  ___        _      ___                       _           
 | _ \___ __| |_   / __|___ _ _  ___ _ _ __ _| |_ ___ _ _ 
 |  _/ _ (_-<  _| | (_ / -_) ' \/ -_) '_/ _` |  _/ _ \ '_|
 |_| \___/__/\__|  \___\___|_||_\___|_| \__,_|\__\___/_|  
                                                         
    '''
    return opening

def getClosing(isPostSuccessful):
    """Creates a string to be printed at the conclusion of the program, which
    varies based on if / how many files were printed successfully and 
    not successfully.
    
    Arguments:
    isPostSuccessful -- Dictionary of blog post names and whether or not they 
                        were successful.
    Return:
    closing -- The string to be printed.
    
    """
    closing = '\n\nUpdate is complete.'
    success = [s for s in isPostSuccessful if isPostSuccessful[s]]
    notSuccess = [ns for ns in isPostSuccessful if not isPostSuccessful[ns]]
    if not isPostSuccessful:
        closing += ('\nNo new posts were found.')
    else:
        if success: 
            closing += ('\nThe following posts were successfully uploaded:\n- ' +
                        ('\n- ').join([s for s in success]))
        else:
            closing += ('\nNo posts were uploaded successfully.')
        if notSuccess: 
            closing += ('The following posts were not successfully uploaded:\n' +
                        ('\n- ').join([ns for ns in notSuccess]))
    closing += '\n\nThank you for using WordPress Post Generator.'
    return closing
    
# ||-------------------------------------------------------------------------||
# ||-Excel Reading-----------------------------------------------------------||
# ||-------------------------------------------------------------------------||

def getNextDaysMeals(day, EXCEL_DIR, SHEET_NAME): 
    """Creates a list of all of the meals and what was eaten for the given
    day. Also passes thisDay from getNextRow() to loopDates().
    
    Arguments:
    day -- The datetime.datetime of the most recent published blog post.
    EXCEL_DIR -- Filename of the Excel file to be read from.
    SHEET_NAME -- Name of the sheet within the Excel file.
    
    Return: (
    meals -- A list of tuples (Meal name, Meal contents),
    thisDay -- The datetime.datetime corresponding to meals.
    ), or (None, None) if there are no unpublished entries in the Excel file.
    
    """
    (row, thisDay) = getNextRow(day, EXCEL_DIR, SHEET_NAME)
    if row == None: return (None, None) 
    meals = [(getMeals(EXCEL_DIR, SHEET_NAME)[ord(meal[1]) - ord('A')-1], 
              meal[3]) for meal in row]
    
    meals = [(meal[0].encode('ascii'),
              meal[1].encode('ascii')) for meal in meals if meal[1]!=None and meal[1].rstrip!='']
    return (meals,thisDay)
    
def getNextRow(day, EXCEL_DIR, SHEET_NAME):
    """Given the date of the most recent blog entry, gets the row in the 
    excel doc corresponding to the next date. Also passes thisDay from 
    getNextRowIndex() to getNextDaysMeals().
    
    Arguments:
    day -- The datetime.datetime of the most recent published blog post.
    EXCEL_DIR -- Filename of the Excel file to be read from.
    SHEET_NAME -- Name of the sheet within the Excel file.
    
    Return: (
    row -- The entire row from the Excel spreadsheet corresponding to the
           next date after day, 
    thisDay -- The datetime.datetime corresponding to meals.
    ), or (None, None) if there are no unpublished entries in the Excel file.
    
    """
    wb = load_workbook(filename = EXCEL_DIR, use_iterators = True)
    sheet_Journal = wb.get_sheet_by_name(name = SHEET_NAME)
    (rowCoordinate, thisDay) = getNextRowIndex(day, EXCEL_DIR, SHEET_NAME)
    if rowCoordinate == -1: return (None,None)
    rangeString = 'B' + str(rowCoordinate) + ':J' + str(rowCoordinate)
    row = [cell for cell in sheet_Journal.iter_rows(rangeString)][0]
    return (row, thisDay)

def getNextRowIndex(day, EXCEL_DIR, SHEET_NAME):
    """Searches through the dates in column A in the Excel file, and returns
    the index of the first date after day.
    
    Arguments:
    day -- The datetime.datetime of the most recent published blog post.
    EXCEL_DIR -- Filename of the Excel file to be read from.
    SHEET_NAME -- Name of the sheet within the Excel file.
    
    Return: (
    index -- the row index of the next unpublished entry in the Excel file, 
    cell.value -- The datetime.datetime corresponding to meals.
    ), or (None, None) if there are no unpublished entries in the Excel file.
    
    """
    wb = load_workbook(EXCEL_DIR)
    sheet_Journal = wb.get_sheet_by_name(SHEET_NAME)
    index = 2
    cell = sheet_Journal.cell("A"+str(index))
    while cell.value != None and day.date() >= cell.value.date(): #
        index += 1
        cell = sheet_Journal.cell("A"+str(index))
    if cell.value == None: return (-1, None)
    return (index, cell.value)

def getMeals(EXCEL_DIR, SHEET_NAME):
    """Returns a list of names of meals in cells B1-I1.
    
    Arguments:
    EXCEL_DIR -- Filename of the Excel file to be read from.
    SHEET_NAME -- Name of the sheet within the Excel file.
    
    Return:
    A list of names of meals in cells B1-I1.
    
    """
    wb = load_workbook(EXCEL_DIR, use_iterators = True)
    sheet_Journal = wb.get_sheet_by_name(SHEET_NAME)
    for line in sheet_Journal.iter_rows('B1:J1'):
        return [cell[3] for cell in line]


# ||-------------------------------------------------------------------------||
# ||-Image Uploading---------------------------------------------------------||
# ||-------------------------------------------------------------------------||

def uploadPictures(wp, date, meals, PHOTOS_DIR):
    """Searches for photos taken on the given date, renames them to the name of
    the meal that they are ('Breakfast.jpeg', 'Dinner.jpeg', etc.), and uploads
    them to WordPress, returning the response data generated by WordPress.
    
    Arguments:
    wp -- The authenticated WordPress client instance.
    date -- The datetime.datetime on which the uploaded photos are taken.
    meals -- A list of tuples (Meal name, Meal contents).
    PHOTOS_DIR -- Directory where photos are stored.
    
    Return:
    uploadedFiles -- A list of dictionaries returned by the WordPress upload 
                     containing the metadata of the uploaded photos.
    """
    uploadedFiles = []
    images = getFileData(date, meals, PHOTOS_DIR)
    for image in images:
        response = wp.call(UploadFile(image))
        uploadedFiles.append(response)
        print '.',
    return uploadedFiles

def getFileData(date, meals, PHOTOS_DIR):
    """Searches through the photos taken on the given date, and return a list
    of dictionaries containing the data needed for upload to WordPress.
    
    Arguments:
    date -- The datetime.datetime on which the uploaded photos are taken.
    meals -- A list of tuples (Meal name, Meal contents).
    PHOTOS_DIR -- Directory where photos are stored.
    
    Return:
    images -- A list of dictionaries containing {'name' : filename
                                                 'bits' : 64-bit encoded binary
                                                          data
                                                 'type' : MIME-type of the file
                                                 }
    """
    mealIndex = 0
    dateRange = getDateRange(date)
    images = []
    for fn in os.listdir(PHOTOS_DIR):
        
        timeStamp = getEXIF(fn, PHOTOS_DIR)['Exif.Image.DateTime']
        if timeStamp > dateRange[0] and timeStamp < dateRange[1]:
            with open(PHOTOS_DIR + fn, 'rb') as binaryImage:
                # for debugging
                # print mealIndex, timeStamp
                imageData = {
                    'name': meals[mealIndex][0]+'.jpeg',
                    'bits': xmlrpc_client.Binary(binaryImage.read()),
                    'type': mimetypes.guess_type(PHOTOS_DIR + fn, 
                                                 strict=True)[0],
                }
            images.append(imageData)
            mealIndex += 1
    return images    
    
def getDateRange(date):
    """Returns a list containing the start and end dates corresponding to the 
    3:30AM-3:30AM 24-hour range surrounding the given date.
    
    Arguments:
    date -- The datetime.datetime on which the date range is centered
    
    Return:
    [startDate, endDate] -- 3:30AM on the given date, 3:30AM on the following
                            day.
    """
    tomorrow = date + datetime.timedelta(1)
    startDate = datetime.datetime(date.year, date.month, date.day, 3,30)
    endDate = datetime.datetime(tomorrow.year, tomorrow.month, 
                                tomorrow.day, 3,30)
    return [startDate, endDate]    

def getEXIF(fn, PHOTOS_DIR):
    """Returns the EXIF data dictionary for the photo of the given filename.
    
    Arguments:
    fn -- The filename of the photo.
    PHOTOS_DIR -- Directory where photos are stored.
        
    Return:
    ret -- A dictionary containing EXIF metadata.
    
    """
    dic = {}
    metadata = pyexiv2.ImageMetadata(PHOTOS_DIR + fn)
    metadata.read()
    for tag in metadata.exif_keys:
        dic[tag] = metadata[tag].value
    return dic


# ||-------------------------------------------------------------------------||
# ||-Wordpress Interfacing---------------------------------------------------||
# ||-------------------------------------------------------------------------||

def getLatestBlogPostDate(wp):
    """Returns the datetime.datetime value of the latest blog entry.
    
    Arguments:
    wp -- The authenticated WordPress client instance.
    
    Return:
    post.date -- datetime.datetime of the most recent blog entry tagged 
                 'Journal Entry', or Nov 1, 2012 if no post is found.
    """
    recentPosts = wp.call(GetPosts(dict(order = 'ASC', post_status='publish')))
    for post in recentPosts:
        for term in post.terms:
            if term.name == 'Journal Entry':
                return post.date
    return datetime.datetime(2012,11,1)
    
def initWordpressConnection(WP_URL, WP_USERNAME):
    """Returns an authenticated WordPress client instance.
    
    Arguments:
    WP_URL -- URL of the WordPress blog.
    WP_USERNAME -- Username associated with the blog.

    Return:
    wp -- The authenticated WordPress client instance.
    
    """
    successful = False
    while not successful:
        successful = True
        wp = Client(WP_URL + 'xmlrpc.php', WP_USERNAME, 
                    getpass.getpass('Enter your WordPress account password:'))
        try:
            username =  wp.call(GetUserInfo())
        except:
            print "Incorrect password."
            successful = False
    print "Connected to "+str(username)+"."
    return wp
    
def postToWordpress(text, thisDay, wp):
    """Publishes the blog post to WordPress.
    
    Arguments:
    text -- HTML text body of the post.
    thisDay -- The date of the post, which will also become the title.
    
    Return:
    post.title -- The title of the WordPress post.
    True / False -- Whether or not the post was successfully posted.
    
    """
    post = WordPressPost()
    post.title = dateToString(thisDay)
    post.content = text
    post.date = (thisDay + 
                datetime.timedelta(0,0,0,0,30,23)) # 6:30pm, EST
    post.post_status = 'publish'
    post.terms_names = {
                        'post_tag': ['Journal Entry', 'Food'],
                        'category': ['Journal']
                        }
    try:
        wp.call(NewPost(post))
    except:
        return (post.title,False)
    return (post.title, True)
    
def dateToString(date):
    """Converts the given date to a string of form 'Weekday, Mmm [D]D,YYYY'.
    
    Arguments:
    date -- Date to be converted to string
    
    Return:
    String representation of date of form 'Weekday, Mmm [D]D,YYYY'.
    
    """
    return (date.strftime('%A, %b ') + 
            date.strftime('%d').lstrip('0') + 
            date.strftime(', %Y'))
    
def getTextBody(meals, uploadedFiles):
    """Forms the HTML text body and returns it as a string.
    
    Arguments:
    meals -- A list of tuples (Meal name, Meal contents).
    uploadedFiles -- A list of dictionaries returned by the WordPress upload 
                     containing the metadata of the uploaded photos.
    
    Return:
    text -- HTML text body of the post.
    
    """
    text = '<h2>[gallery orderby="ID" order="ASC" link="file" include="'
    text += (', ').join([response['id'] 
                         for response in uploadedFiles]) + '"]</h2>\n'
    for meal in meals:
        text += '<h2>' + meal[0] + '</h2>\n<ul>\n<li>'
        text += ('</li>\n<li>').join([food.strip() 
                                      for food in meal[1].split(';')])
        text += '</li>\n</ul>\n'
    return text

# ||----------||
# ||-Run Main-||
# ||----------||
if __name__ == '__main__': main()